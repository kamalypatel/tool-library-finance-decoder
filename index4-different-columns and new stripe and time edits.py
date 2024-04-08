from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
import datetime


right_aligned_text = Alignment(horizontal="right")

filename = "End.xlsx"
workbookwrite = Workbook()
sheetwrite = workbookwrite.active

sheetwrite["A1"] = 'Sales receipt no'
sheetwrite["B1"] = 'Ref no'
sheetwrite["C1"] = 'Customer'
sheetwrite["D1"] = 'Sales Receipt Date'
sheetwrite["E1"] = 'Payment Method'
sheetwrite["F1"] = 'Deposit to'
sheetwrite["G1"] = 'Location of Sale'
sheetwrite["H1"] = 'Total'
sheetwrite["I1"] = 'Amount'
sheetwrite["J1"] = 'Description'
sheetwrite["K1"] = 'Product/service'
sheetwrite["L1"] = 'Class'
sheetwrite["M1"] = 'Location'

workbookread = load_workbook(filename="Start.xlsx")
sheetread = workbookread.active

alpha = sheetread.title
i = 2
k = 2
while (i <= 5000):
    if sheetread["N"+str(i)].value == "Credit/Debit (Stripe)" or sheetread["N"+str(i)].value == "Cash":
        #need to add the Sales Reciept number manually
        value = sheetread["A" + str(i)].value
        sheetwrite["B"+str(k)] = value

        value = sheetread["B" + str(i)].value
        sheetwrite["C"+str(k)] = value

        #date handling
        value = sheetread["Z" + str(i)].value
        if isinstance(value, datetime.datetime):
            value = value.strftime("%d/%m/%Y")  # Formats the date as 'YYYY-MM-DD'
        else:
            if isinstance(value, str) and len(value) >= 10:
                if value[1] == "/":
                    value = "0" + value
                value = value[:10]


        sheetwrite["D"+str(k)] = value
        
        #handling payment type
        value = sheetread["N" + str(i)].value
        if value == "Credit/Debit (Stripe)":
            writer = "Stripe"
            writertwo = "Stripe clearing account"
        else:
            writer = "Cash"
            writertwo = "cash on hand"
        
        sheetwrite["E"+str(k)] = writer
        sheetwrite["F"+str(k)] = writertwo

        value = "5 W. Northrup Place Buffalo NY 14214 US"
        sheetwrite["G"+str(k)] = value

        value = sheetread["O" + str(i)].value
        sheetwrite["H"+str(k)] = value
        sheetwrite["I"+str(k)] = value
        sheetwrite["H" + str(k)].alignment = right_aligned_text
        sheetwrite["I" + str(k)].alignment = right_aligned_text


        #handling the type of transaction and category
        j = i-1
        check = sheetread["AA" + str(j)].value
        if check == "Membership Fee":
            checktwo = sheetread["AF" + str(j)].value
            sheetwrite["J" + str(k)] = checktwo
            if checktwo == "Tool Belt" or checktwo == "Tool Box" or checktwo == "Wheelbarrow":
                checkthree = "Memberships:" + checktwo + " Membership"
                sheetwrite["K" + str(k)] = checkthree
            else:
                checkthree = "Memberships:" + checktwo
                sheetwrite["K" + str(k)] = checkthree
        

        #handling the category
        if check == "Loan Fee":
            checktwo = "supertool"
            sheetwrite["J" + str(k)] = checktwo
            checkthree = "Memberships:Supertool Rental"
            sheetwrite["K" + str(k)] = checkthree

        if check == "Late Fee":
            checktwo = "Late Fee"
            sheetwrite["J" + str(k)] = checktwo
            checkthree = "Donations:Late Fees"
            sheetwrite["K" + str(k)] = checkthree

        if check == "Donation":
            checktwo = "Donation"
            sheetwrite["J" + str(k)] = checktwo
            checkthree = "Donations:Individual Contribution"
            sheetwrite["K" + str(k)] = checkthree  
        

        if check == "NYS & Erie County Sales Tax":
            checkagain = sheetread["AA" + str(j-1)].value
            
            if checkagain == "Cart Item Purchase":
                checktwo = "Cart Item Purchase"
                sheetwrite["J" + str(k)] = checktwo
                checkthree = "Merchandise:Cart Items for sale"
                sheetwrite["K" + str(k)] = checkthree
                #doing the amount
                value = sheetread["O" + str(i)].value
                sheetwrite["H"+str(k)] = value
                sheetwrite["I"+str(k)] = round(float(value) - (float(value) * 0.0875), 2)
                sheetwrite["H" + str(k)].alignment = right_aligned_text
                sheetwrite["I" + str(k)].alignment = right_aligned_text
            elif checkagain == "Consumables":
                checktwo = "Consumables"
                sheetwrite["J" + str(k)] = checktwo
                checkthree = "Merchandise:Consumables"
                sheetwrite["K" + str(k)] = checkthree
                #doing the amount
                value = sheetread["O" + str(i)].value
                sheetwrite["H"+str(k)] = value
                sheetwrite["I"+str(k)] = round(float(value) - (float(value) * 0.0875), 2)
                sheetwrite["H" + str(k)].alignment = right_aligned_text
                sheetwrite["I" + str(k)].alignment = right_aligned_text
            else:
                continue


        #for times that there is no NYS Sales in the first spot
        checkagain = sheetread["AA" + str(j)].value
        if checkagain == "Cart Item Purchase":
            checktwo = "Cart Item Purchase"
            sheetwrite["J" + str(k)] = checktwo
            checkthree = "Merchandise:Cart Items for sale"
            sheetwrite["K" + str(k)] = checkthree
            #doing the amount
            value = sheetread["O" + str(i)].value
            sheetwrite["H"+str(k)] = value
            sheetwrite["I"+str(k)] = round(float(value) - (float(value) * 0.0875), 2)
            sheetwrite["H" + str(k)].alignment = right_aligned_text
            sheetwrite["I" + str(k)].alignment = right_aligned_text

        if check == "Consumables":
            checktwo = "Consumables"
            sheetwrite["J" + str(k)] = checktwo
            checkthree = "Merchandise:Consumables"
            sheetwrite["K" + str(k)] = checkthree
             #doing the amount
            value = sheetread["O" + str(i)].value
            sheetwrite["H"+str(k)] = value
            sheetwrite["I"+str(k)] = round(float(value) - (float(value) * 0.0875), 2)
            sheetwrite["H" + str(k)].alignment = right_aligned_text
            sheetwrite["I" + str(k)].alignment = right_aligned_text
        elif check == "Check Out (Renewal)" or check == "Check Out" or check == "":
            value = sheetread["O" + str(i)].value
            sheetwrite["H"+str(k)] = value
            sheetwrite["I"+str(k)] = ""
            sheetwrite["H" + str(k)].alignment = right_aligned_text
            sheetwrite["I" + str(k)].alignment = right_aligned_text

        #if check != "Consumables" or check != "Cart Item Purchase" or check != "NYS & Erie County Sales Tax" or check != "Donation" or check != "Late Fee" or check != "Loan Fee" or check != "Membership Fee":
        #    sheetwrite["I"+str(k)] = ""

        value = "Programs:Tool Lending Library"
        sheetwrite["L"+str(k)] = value

        value = "1 WITHOUT DONOR RESTRICTIONS"
        sheetwrite["M"+str(k)] = value
        k += 1

    i += 1

workbookwrite.save(filename = filename)